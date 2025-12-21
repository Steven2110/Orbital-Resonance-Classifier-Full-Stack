from rest_framework import status
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from django.conf import settings
from django.http import HttpResponse
from .models import TextFile, GeneratedImage, Prediction
from .serializers import TextFileSerializer, GeneratedImageSerializer, PredictionSerializer
import os
import random
import zipfile
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image, ImageDraw, ImageFont
import tempfile


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_files(request):
    """
    Upload multiple text files
    POST /api/upload/
    """
    files = request.FILES.getlist('files')
    
    if not files:
        return Response(
            {'error': 'No files provided'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    uploaded_files = []
    for file in files:
        # Validate file type
        if not file.name.endswith('.txt'):
            continue
            
        text_file = TextFile.objects.create(
            file=file,
            filename=file.name
        )
        uploaded_files.append(text_file)
    
    serializer = TextFileSerializer(uploaded_files, many=True)
    return Response({
        'message': f'{len(uploaded_files)} files uploaded successfully',
        'files': serializer.data
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
def list_files(request):
    """
    List all uploaded text files
    GET /api/files/
    """
    files = TextFile.objects.all().order_by('-uploaded_at')
    serializer = TextFileSerializer(files, many=True)
    return Response(serializer.data)


@api_view(['DELETE'])
def delete_file(request, file_id):
    """
    Delete a specific text file
    DELETE /api/files/<id>/
    """
    try:
        text_file = TextFile.objects.get(id=file_id)
        # Delete the actual file
        if text_file.file and os.path.exists(text_file.file.path):
            os.remove(text_file.file.path)
        text_file.delete()
        return Response({'message': 'File deleted successfully'}, status=status.HTTP_200_OK)
    except TextFile.DoesNotExist:
        return Response({'error': 'File not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
def generate_image(request):
    """
    Generate image from selected text files
    POST /api/generate-image/
    Body: { "file_ids": [1, 2, 3] }
    
    This endpoint will call your custom image generation function.
    You should import and call your function here.
    """
    file_ids = request.data.get('file_ids', [])
    
    if not file_ids:
        return Response(
            {'error': 'No file IDs provided'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Get the text files
    text_files = TextFile.objects.filter(id__in=file_ids)
    
    if not text_files.exists():
        return Response(
            {'error': 'No valid files found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    try:
        # TODO: Import and call your image generation function here
        # Example:
        # from your_module import generate_image_function
        # image_path = generate_image_function(text_files)
        
        # For now, create a placeholder
        generated_image = GeneratedImage.objects.create()
        generated_image.text_files.set(text_files)
        
        # You should set the actual generated image file here:
        # generated_image.image = image_path
        # generated_image.save()
        
        serializer = GeneratedImageSerializer(generated_image)
        return Response({
            'message': 'Image generation initiated',
            'image': serializer.data,
            'note': 'Replace this with your actual image generation function'
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response(
            {'error': f'Image generation failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def list_images(request):
    """
    List all generated images
    GET /api/images/
    """
    images = GeneratedImage.objects.all().order_by('-created_at')
    serializer = GeneratedImageSerializer(images, many=True)
    return Response(serializer.data)


@api_view(['POST'])
def predict_from_image(request, image_id):
    """
    Run ML prediction on a generated image
    POST /api/images/<id>/predict/
    
    This endpoint will call your custom ML prediction function.
    You should import and call your function here.
    """
    try:
        generated_image = GeneratedImage.objects.get(id=image_id)
    except GeneratedImage.DoesNotExist:
        return Response(
            {'error': 'Image not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    try:
        # TODO: Import and call your ML prediction function here
        # Example:
        # from your_module import predict_function
        # prediction_result = predict_function(generated_image.image.path)
        
        # For now, create a placeholder prediction
        prediction = Prediction.objects.create(
            image=generated_image,
            prediction_result={
                'status': 'placeholder',
                'message': 'Replace this with your actual ML prediction function'
            },
            confidence=0.0
        )
        
        # You should set actual prediction results here:
        # prediction.prediction_result = prediction_result
        # prediction.confidence = confidence_score
        # prediction.save()
        
        serializer = PredictionSerializer(prediction)
        return Response({
            'message': 'Prediction completed',
            'prediction': serializer.data,
            'note': 'Replace this with your actual ML prediction function'
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response(
            {'error': f'Prediction failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def list_predictions(request):
    """
    List all predictions
    GET /api/predictions/
    """
    predictions = Prediction.objects.all().order_by('-created_at')
    serializer = PredictionSerializer(predictions, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def get_prediction(request, prediction_id):
    """
    Get a specific prediction
    GET /api/predictions/<id>/
    """
    try:
        prediction = Prediction.objects.get(id=prediction_id)
        serializer = PredictionSerializer(prediction)
        return Response(serializer.data)
    except Prediction.DoesNotExist:
        return Response(
            {'error': 'Prediction not found'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_and_predict(request):
    """
    Upload files and immediately return predictions
    POST /api/upload-and-predict/
    
    Returns predictions in format:
    {
        "predictions": [
            {
                "filename": "file1.txt",
                "phi1": 0.85,
                "phi2": 0.92,
                "phi3": 0.78,
                "phi4": 0.88,
                "phi5": 0.95
            }
        ]
    }
    
    TODO: Replace the mock prediction logic with your actual ML model
    """
    files = request.FILES.getlist('files')
    
    if not files:
        return Response(
            {'error': 'No files provided'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    predictions = []
    
    for file in files:
        # Validate file type
        if not file.name.endswith('.txt'):
            continue
        
        # Save file to database
        text_file = TextFile.objects.create(
            file=file,
            filename=file.name
        )
        
        try:
            # Use ML model to predict from generated images
            from .ml_predictor import predict_from_images
            
            # Generate images and run predictions
            # Returns: {'phi1': 0, 'phi2': 1, 'phi3': 2, 'phi4': 0, 'phi5': 1}
            ml_predictions = predict_from_images(text_file.file.path)
            
            # Combine filename with predictions
            phi_values = {
                'filename': file.name,
                **ml_predictions  # Merge ML predictions
            }
            
            predictions.append(phi_values)
            
        except Exception as e:
            # If prediction fails for a file, include error
            error_msg = str(e)
            print(f"Error processing {file.name}: {error_msg}")
            predictions.append({
                'filename': file.name,
                'error': error_msg
            })
    
    if not predictions:
        return Response(
            {'error': 'No valid text files provided'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Check if any predictions have errors
    has_errors = any('error' in pred for pred in predictions)
    
    if has_errors:
        # If there are errors, return 500 status
        error_files = [pred['filename'] for pred in predictions if 'error' in pred]
        return Response({
            'error': f'Error processing files: {", ".join(error_files)}',
            'details': predictions
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    return Response({
        'message': f'{len(predictions)} files processed successfully',
        'predictions': predictions
    }, status=status.HTTP_200_OK)


def generate_scatter_plot_image(text_file_path, phi_index):
    """
    Generate scatter plot image from text file data
    
    Args:
        text_file_path: Path to the text file with data
        phi_index: Which Phi column to plot (1-5)
    
    Returns:
        PIL Image object (224x224 pixels, white background)
    """
    import matplotlib
    matplotlib.use('Agg')  # Use non-interactive backend
    import matplotlib.pyplot as plt
    import numpy as np
    
    # Configuration
    DELIMITER = '\t'
    DPI = 100
    FIGURE_SIZE_INCHES = 2.24
    
    try:
        # 1. Load the data (Column 0 for X, Column phi_index for Y)
        data = np.loadtxt(text_file_path, delimiter=DELIMITER)
        x_data = data[:, 0]
        y_data = data[:, phi_index]
        
        # 2. Create the Figure with a white background
        fig = plt.figure(
            figsize=(FIGURE_SIZE_INCHES, FIGURE_SIZE_INCHES),
            frameon=False
        )
        fig.set_facecolor('white')
        
        # 3. Create the Axes object to cover the entire figure area
        ax = fig.add_axes([0, 0, 1, 1])
        ax.set_axis_off()
        ax.set_facecolor('white')
        
        # 4. Plot the data as a SCATTER of black points
        ax.scatter(x_data, y_data, color='black', marker='o', s=1)
        
        # 5. Set limits to exact data range (no buffer/padding)
        x_min, x_max = x_data.min(), x_data.max()
        y_min, y_max = y_data.min(), y_data.max()
        ax.set_xlim(x_min, x_max)
        ax.set_ylim(y_min, y_max)
        
        # 6. Save to BytesIO buffer
        buf = io.BytesIO()
        plt.savefig(
            buf,
            format='png',
            dpi=DPI,
            bbox_inches='tight',
            pad_inches=0,
            facecolor='white',
            edgecolor='white',
            transparent=False
        )
        plt.close(fig)
        buf.seek(0)
        
        # 7. Convert to RGB and ensure white background
        img = Image.open(buf)
        
        # Create a white background image
        if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
            img = img.convert('RGBA')
            white_bg = Image.new('RGB', img.size, (255, 255, 255))
            white_bg.paste(img, mask=img.split()[3])
        else:
            white_bg = img.convert('RGB')
        
        # Resize to 224x224 without adding white edges
        resized_img = white_bg.resize((224, 224), Image.Resampling.LANCZOS)
        
        return resized_img
        
    except Exception as e:
        print(f"Error generating scatter plot for Phi {phi_index}: {e}")
        # Return a blank white image on error
        return Image.new('RGB', (224, 224), color=(255, 255, 255))


@api_view(['POST'])
def download_results(request):
    """
    Download results as a zip file containing:
    - images/ folder with generated images
    - results.xlsx with prediction table
    
    POST /api/download-results/
    Body: { "predictions": [...] }
    """
    predictions = request.data.get('predictions', [])
    
    if not predictions:
        return Response(
            {'error': 'No predictions provided'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Create a BytesIO buffer for the zip file
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Prediction Results"
        
        # Style the header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        headers = ['File Name', 'Φ1', 'Φ2', 'Φ3', 'Φ4', 'Φ5']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Category mapping
        categories = {0: 'Circulation', 1: 'Libration/Circulation', 2: 'Libration'}
        
        # Write data
        for row_idx, pred in enumerate(predictions, 2):
            ws.cell(row=row_idx, column=1, value=pred.get('filename', ''))
            ws.cell(row=row_idx, column=2, value=categories.get(pred.get('phi1'), str(pred.get('phi1'))))
            ws.cell(row=row_idx, column=3, value=categories.get(pred.get('phi2'), str(pred.get('phi2'))))
            ws.cell(row=row_idx, column=4, value=categories.get(pred.get('phi3'), str(pred.get('phi3'))))
            ws.cell(row=row_idx, column=5, value=categories.get(pred.get('phi4'), str(pred.get('phi4'))))
            ws.cell(row=row_idx, column=6, value=categories.get(pred.get('phi5'), str(pred.get('phi5'))))
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save Excel to buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Add Excel file to zip
        zip_file.writestr('results.xlsx', excel_buffer.read())
        
        # Generate and add images (5 images per file, one for each Φ)
        for pred in predictions:
            filename = pred.get('filename', 'unknown')
            
            # Find the uploaded text file
            try:
                text_file = TextFile.objects.filter(filename=filename).first()
                if not text_file or not text_file.file:
                    print(f"Text file not found for {filename}")
                    continue
                
                text_file_path = text_file.file.path
                base_filename = filename.replace('.txt', '')
                
                # Generate 5 scatter plot images (one for each Φ column)
                for phi_index in range(1, 6):
                    try:
                        # Generate scatter plot image
                        img = generate_scatter_plot_image(text_file_path, phi_index)
                        
                        # Save image to buffer as JPEG
                        img_buffer = io.BytesIO()
                        img.save(img_buffer, format='JPEG', quality=95)
                        img_buffer.seek(0)
                        
                        # Add image to zip in images folder
                        image_filename = f'{base_filename}_Ф{phi_index}.jpg'
                        zip_file.writestr(f'images/{image_filename}', img_buffer.read())
                        
                    except Exception as e:
                        print(f"Error generating Φ{phi_index} image for {filename}: {e}")
                
            except Exception as e:
                print(f"Error processing {filename}: {e}")
    
    # Prepare the zip file for download
    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="prediction_results.zip"'
    
    return response
